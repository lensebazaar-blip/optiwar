#!/usr/bin/env python3
"""Import a contact-lens export into products + profile + matrix.

Dry run by default. Nothing is written until a person types ``--apply``, and
what is written is decided entirely by ``cl_import.parse()``, so a rejection is
provable in a test without a database.

    # what would happen
    python3 scripts/import_contact_lenses.py \
        --products products.csv --variants variants.csv

    # do it
    python3 scripts/import_contact_lenses.py \
        --products products.csv --variants variants.csv --by sudhanshu --apply

CSV (or TSV) is what this reads. An .xlsx is read only if openpyxl happens to be
installed — the export arrives as a workbook, and one sheet saved as CSV is a
smaller dependency than a spreadsheet parser in the deployment.

Per product, in one transaction:

    products                    the commercial record, CONTACT_LENS, .com only
    contact_lens_products       brand, manufacturer, modality, pack, minimums
    contact_lens_param_rules    the stated values, when param_mode is RULES
    contact_lens_variants       the matrix, when param_mode is MATRIX
    contact_lens_images         the primary image, or every approved view when
                                ``--images`` names the product's image recipe

A product is stated in one shape or the other, never both: ``--rules`` carries
the parameter sheet and ``--variants`` the combinations sheet, and the product
row's ``param_mode`` says which of them applies to it.

Idempotent on (source_system, source_ref): re-running the same export updates
the same products instead of making new ones. Nothing is ever deleted — a
combination the manufacturer withdraws becomes ``available = 0``, so an order
that referenced it stays explicable.

``merchant_enabled`` stays 0. An imported lens is in the database and on no
surface until somebody releases it.

Connection comes from MYSQL_HOST/MYSQL_USER/MYSQL_PASSWORD/MYSQL_DB, the same
variables gunicorn runs with; on the box they are the service's Environment=
lines, so export them from ``systemctl show gunicorn -p Environment`` before
running it there.
"""
import argparse
import csv
import datetime
import os
import sys

import pymysql

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import cl_import  # noqa: E402
import contact_lens  # noqa: E402
import image_pipeline  # noqa: E402

# The .com-only launch. Site eligibility lives on products because one function
# in catalogue.py decides it for every vertical and every surface.
SELL_ON = {"sell_on_com": 1, "sell_on_in": 0}


def read_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        return read_workbook(path)
    with open(path, newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        delimiter = "\t" if "\t" in sample.splitlines()[0] else ","
        return [dict(row) for row in csv.DictReader(fh, delimiter=delimiter)]


def read_workbook(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("%s is a workbook and openpyxl is not installed — save the "
                 "sheet as CSV, or pip install openpyxl" % path)
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c or "").strip() for c in next(rows)]
    return [dict(zip(header, row)) for row in rows]


SOCKETS = ("/var/lib/mysql/mysql.sock", "/var/run/mysqld/mysqld.sock")


def connect():
    """The connection gunicorn has: MySQLdb reads ``localhost`` as the unix
    socket, and a box whose ``localhost`` resolves to ::1 while the server
    listens on IPv4 refuses the TCP form, so the socket is used when it is
    there."""
    host = os.environ.get("MYSQL_HOST", "localhost")
    options = {}
    socket_path = os.environ.get("MYSQL_UNIX_SOCKET") or next(
        (p for p in SOCKETS if host == "localhost" and os.path.exists(p)),
        None)
    if socket_path:
        options["unix_socket"] = socket_path
    return pymysql.connect(
        host=host,
        user=os.environ.get("MYSQL_USER", ""),
        password=os.environ.get("MYSQL_PASSWORD", ""),
        database=os.environ.get("MYSQL_DB", ""),
        cursorclass=pymysql.cursors.DictCursor, autocommit=False, **options)


def existing(cursor, product):
    cursor.execute("SELECT product_id FROM contact_lens_products"
                   " WHERE source_system = %s AND source_ref = %s",
                   (product["source_system"], product["source_ref"]))
    row = cursor.fetchone()
    return row["product_id"] if row else None


def product_code(product):
    """Our internal offer id. Never sent as the manufacturer's identifier."""
    return ("CL-" + product["source_ref"].upper())[:20]


def slug(product):
    name = product["product_name"]
    text = name if name.lower().startswith(product["brand"].lower()) else (
        "%s %s" % (product["brand"], name))
    keep = [c.lower() if c.isalnum() else "-" for c in text]
    return "-".join("".join(keep).split("-")[:12]).strip("-")[:180]


def in_rupees(amount, rate):
    """EUR -> INR at a stated rate, or None when no rate was supplied.

    EUR is canonical. A rupee price is derived from it once, at a rate the run
    was given and records; it is never re-derived from a previous conversion,
    which is how a price drifts every time somebody re-imports.
    """
    if not rate or amount in (None, ""):
        return None
    # products.product_price is whole rupees, so the rounding is done here
    # where it is stated rather than by the column on the way in.
    return int(round(float(amount) * float(rate)))


def upsert_product(cursor, product, product_id, rate=None):
    fields = {
        "product_code": product_code(product),
        "product_name": product["product_name"],
        "product_details": product["product_details"],
        "product_price_eur": product["price_eur"],
        "product_special_price_eur": product["special_price_eur"],
        "product_image": product["image_url"],
        "product_slug": slug(product),
        "product_vertical": contact_lens.VERTICAL,
        "product_status": "ACTIVE",
    }
    fields.update(SELL_ON)
    rupees = in_rupees(product["price_eur"], rate)
    if rupees is not None:
        fields["product_price"] = rupees
        fields["product_special_price"] = in_rupees(
            product["special_price_eur"] or product["price_eur"], rate)
    if product_id:
        assignments = ", ".join("%s = %%s" % k for k in fields)
        cursor.execute("UPDATE products SET %s WHERE product_id = %%s"
                       % assignments,
                       tuple(fields.values()) + (product_id,))
        return product_id
    columns = ", ".join(fields)
    marks = ", ".join(["%s"] * len(fields))
    cursor.execute("INSERT INTO products (%s) VALUES (%s)" % (columns, marks),
                   tuple(fields.values()))
    return cursor.lastrowid


def upsert_profile(cursor, product, product_id, rate=None):
    fields = {
        "product_id": product_id,
        "brand": product["brand"],
        "manufacturer": product["manufacturer"],
        "source_manufacturer": product["source_manufacturer"] or None,
        "param_mode": product["param_mode"],
        "param_source": product["param_source"] or None,
        "min_boxes_single_eye": product["min_boxes_single_eye"],
        "min_boxes_both_per_eye": product["min_boxes_both_per_eye"],
        "min_order_model": product["min_order_model"] or None,
        "gtin": product["gtin"] or None,
        "manufacturer_mpn": product["manufacturer_mpn"] or None,
        "modality": product["modality"],
        "lens_type": product["lens_type"],
        "pack_quantity": product["pack_quantity"],
        "material": product["material"] or None,
        "water_content": product["water_content"],
        "replacement_days": product["replacement_days"],
        "availability": product["availability"],
        "lead_time_days": product["lead_time_days"],
        "source_system": product["source_system"],
        "source_ref": product["source_ref"],
        "imported_at": datetime.datetime.now(),
        "eur_inr_rate": rate,
        "eur_inr_rate_at": datetime.datetime.now() if rate else None,
    }
    columns = ", ".join(fields)
    marks = ", ".join(["%s"] * len(fields))
    # merchant_enabled is absent on purpose: an update must not re-release a
    # lens somebody withdrew, and an insert takes the column's default of 0.
    # The conversion metadata is held back for the same reason upsert_product
    # leaves the rupee prices alone when no rate was given: the recorded rate
    # describes the rupee price that is still there, and blanking it would
    # leave a converted price nothing accounts for.
    kept = {"product_id"} if rate else {"product_id", "eur_inr_rate",
                                        "eur_inr_rate_at"}
    updates = ", ".join("%s = VALUES(%s)" % (k, k) for k in fields
                        if k not in kept)
    cursor.execute("INSERT INTO contact_lens_products (%s) VALUES (%s)"
                   " ON DUPLICATE KEY UPDATE %s" % (columns, marks, updates),
                   tuple(fields.values()))


def upsert_variants(cursor, product, product_id):
    """Upsert every stated combination; withdraw the ones no longer stated.

    Withdrawal is ``available = 0`` rather than a DELETE, because an order line
    that pointed at a combination must remain readable after the manufacturer
    stops making it.
    """
    stated = set()
    for variant in product["variants"]:
        cursor.execute(
            "INSERT INTO contact_lens_variants (product_id, sph, cyl, axis,"
            " add_power, base_curve, diameter, color_code, color_name,"
            " available) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            " ON DUPLICATE KEY UPDATE color_name = VALUES(color_name),"
            " available = VALUES(available)",
            (product_id, variant["sph"], variant["cyl"], variant["axis"],
             variant["add_power"], variant["base_curve"], variant["diameter"],
             variant["color_code"], variant["color_name"] or None,
             variant["available"]))
        stated.add(cl_import.variant_signature(variant))
    cursor.execute("SELECT variant_id, variant_sig FROM contact_lens_variants"
                   " WHERE product_id = %s AND available = 1", (product_id,))
    withdrawn = [r["variant_id"] for r in (cursor.fetchall() or ())
                 if r["variant_sig"] not in stated]
    for variant_id in withdrawn:
        cursor.execute("UPDATE contact_lens_variants SET available = 0"
                       " WHERE variant_id = %s", (variant_id,))
    return len(stated), len(withdrawn)


def upsert_rules(cursor, product, product_id):
    """Upsert every stated value; withdraw the ones no longer stated.

    Withdrawal is ``available = 0`` for the same reason a combination's is: an
    order that was placed on a power the supplier has dropped stays readable.
    """
    stated = set()
    for order, rule in enumerate(product["rules"]):
        cursor.execute(
            "INSERT INTO contact_lens_param_rules (product_id, parameter,"
            " value, label, sort_order, available)"
            " VALUES (%s, %s, %s, %s, %s, %s)"
            " ON DUPLICATE KEY UPDATE label = VALUES(label),"
            " sort_order = VALUES(sort_order), available = VALUES(available)",
            (product_id, rule["parameter"], rule["value"], rule["label"],
             order, rule["available"]))
        stated.add((rule["parameter"], rule["value"]))
    cursor.execute("SELECT rule_id, parameter, value FROM"
                   " contact_lens_param_rules"
                   " WHERE product_id = %s AND available = 1", (product_id,))
    withdrawn = [r["rule_id"] for r in (cursor.fetchall() or ())
                 if (r["parameter"], r["value"]) not in stated]
    for rule_id in withdrawn:
        cursor.execute("UPDATE contact_lens_param_rules SET available = 0"
                       " WHERE rule_id = %s", (rule_id,))
    return len(stated), len(withdrawn)


def upsert_image(cursor, product, product_id):
    cursor.execute("SELECT image_id FROM contact_lens_images"
                   " WHERE product_id = %s AND image_url = %s",
                   (product_id, product["image_url"]))
    if cursor.fetchone():
        return
    cursor.execute("INSERT INTO contact_lens_images (product_id, color_code,"
                   " image_url, image_type, sort_order)"
                   " VALUES (%s, NULL, %s, 'PRIMARY', 0)",
                   (product_id, product["image_url"]))


def upsert_views(cursor, recipe, product_id):
    """One row per approved view of the recipe, keyed on the view code.

    A row is matched by ``view_code`` or, for imagery loaded before views were
    recorded, by ``image_url``, and updated in place; nothing is deleted. The
    recipe is the only source of what the images are, so the gallery, the
    feed and the sitemap cannot disagree with what was photographed. A view
    the recipe no longer names is marked WITHDRAWN, which every reader skips,
    so a withdrawn photograph stops being published without a row being lost.
    """
    written = 0
    records = image_pipeline.image_records(recipe)
    codes = [r["code"] for r in records]
    cursor.execute("UPDATE contact_lens_images SET image_type = 'WITHDRAWN',"
                   " gmc_eligible = 0 WHERE product_id = %%s AND (color_code"
                   " IS NULL OR color_code = '') AND view_code IS NOT NULL"
                   " AND view_code NOT IN (%s)"
                   % ", ".join(["%s"] * len(codes)),
                   (product_id,) + tuple(codes))
    for record in records:
        fields = {
            "image_url": record["path"],
            "image_type": "PRIMARY" if record["is_primary"] else "GALLERY",
            "sort_order": record["position"],
            "view_code": record["code"],
            "view_name": record["view"],
            "alt_text": record["alt"],
            "gmc_eligible": 1 if record["gmc"] else 0,
        }
        cursor.execute("SELECT image_id FROM contact_lens_images"
                       " WHERE product_id = %s AND (color_code IS NULL OR"
                       " color_code = '') AND (view_code = %s OR"
                       " image_url = %s) ORDER BY image_id LIMIT 1",
                       (product_id, record["code"], record["path"]))
        row = cursor.fetchone()
        if row:
            assignments = ", ".join("%s = %%s" % k for k in fields)
            cursor.execute("UPDATE contact_lens_images SET %s"
                           " WHERE image_id = %%s" % assignments,
                           tuple(fields.values()) + (row["image_id"],))
        else:
            columns = ", ".join(["product_id", "color_code"] + list(fields))
            marks = ", ".join(["%s", "NULL"] + ["%s"] * len(fields))
            cursor.execute("INSERT INTO contact_lens_images (%s) VALUES (%s)"
                           % (columns, marks),
                           (product_id,) + tuple(fields.values()))
        written += 1
    return written


def recipe_for(recipes, product):
    """The image recipe whose product code is this source_ref, or None."""
    return recipes.get(product["source_ref"].upper())


def image_url_is_primary(recipe, image_url):
    """True only for the recipe's primary view: the product's lead image is
    the hero, never a gallery view that happens to be in the recipe."""
    return any(r["is_primary"] and r["path"] == image_url
               for r in image_pipeline.image_records(recipe))


def withdraw_all(cursor, table, product_id):
    """Withdraw whatever the shape a product no longer uses still offers.

    A lens states what may be ordered in one shape or the other. If it changes
    shape, the rows of the shape it left are still marked available, and the
    storefront would have two answers to the same question. They are withdrawn,
    not deleted, so an order placed against one stays readable.
    """
    cursor.execute("UPDATE %s SET available = 0"
                   " WHERE product_id = %%s AND available = 1" % table,
                   (product_id,))
    return cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else 0


def import_one(cursor, product, rate=None, recipe=None):
    """One product and everything it states, inside the caller's transaction."""
    product_id = existing(cursor, product)
    product_id = upsert_product(cursor, product, product_id, rate)
    upsert_profile(cursor, product, product_id, rate)
    if product["param_mode"] == cl_import.PARAM_MODE_RULES:
        written, withdrawn = upsert_rules(cursor, product, product_id)
        withdrawn += withdraw_all(cursor, "contact_lens_variants", product_id)
    else:
        written, withdrawn = upsert_variants(cursor, product, product_id)
        withdrawn += withdraw_all(cursor, "contact_lens_param_rules",
                                  product_id)
    if recipe:
        upsert_views(cursor, recipe, product_id)
    else:
        upsert_image(cursor, product, product_id)
    return product_id, written, withdrawn


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--products", required=True, help="product sheet (CSV/TSV)")
    ap.add_argument("--variants", default=None,
                    help="prescription combinations, for a MATRIX product")
    ap.add_argument("--rules", default=None,
                    help="selectable parameter values, for a RULES product")
    ap.add_argument("--only", default=None,
                    help="comma-separated source_refs — the pilot import")
    ap.add_argument("--by", default=None, help="who authorised it")
    ap.add_argument("--eur-inr-rate", type=float, default=None,
                    help="derive the INR columns from the EUR price at this "
                         "rate, and record the rate against the product")
    ap.add_argument("--images", action="append", default=[],
                    help="an image recipe (image_recipes/<CODE>.json); its "
                         "approved views become the product's image rows. "
                         "Repeatable. Matched to the product whose source_ref "
                         "is the recipe's product code")
    ap.add_argument("--apply", action="store_true",
                    help="write; otherwise dry run")
    args = ap.parse_args()

    recipes = {}
    for path in args.images:
        recipe = image_pipeline.load_recipe(path)
        recipes[recipe["product"].upper()] = recipe

    if not (args.variants or args.rules):
        sys.exit("--variants or --rules is required: a product row on its own "
                 "does not say what may be ordered")
    products, errors = cl_import.parse(
        read_rows(args.products),
        read_rows(args.variants) if args.variants else [],
        read_rows(args.rules) if args.rules else [])
    if args.only:
        wanted = {r.strip() for r in args.only.split(",") if r.strip()}
        unknown = wanted - {p["source_ref"] for p in products}
        if unknown:
            sys.exit("not importable (rejected or absent): %s"
                     % ", ".join(sorted(unknown)))
        products = [p for p in products if p["source_ref"] in wanted]

    print(cl_import.report(products, errors))
    for product in products:
        recipe = recipe_for(recipes, product)
        if recipe:
            records = image_pipeline.image_records(recipe)
            if not image_url_is_primary(recipe, product["image_url"]):
                sys.exit("%s: image_url %s is not the primary view of recipe %s"
                         % (product["source_ref"], product["image_url"],
                            recipe["product"]))
            print("  images %-14s %d view(s) from recipe, %d GMC-eligible"
                  % (product["source_ref"], len(records),
                     sum(1 for r in records if r["gmc"])))
            for level, text in image_pipeline.qa_warnings(recipe, records):
                print("         %-5s %s" % (level, text))
    unmatched = set(recipes) - {p["source_ref"].upper() for p in products}
    if unmatched:
        sys.exit("--images recipe(s) match no importable product: %s"
                 % ", ".join(sorted(unmatched)))
    if not products:
        sys.exit(1 if errors else 0)
    if not args.apply:
        print("\nDRY RUN — nothing written. Re-run with --by and --apply.")
        return
    if not args.by:
        sys.exit("--by is required to write")

    db = connect()
    cursor = db.cursor()
    contact_lens.ensure_schema(cursor)
    db.commit()

    failures = []
    for product in products:
        # One transaction per product: a product whose matrix fails leaves
        # nothing behind, and the products that already succeeded stay.
        try:
            product_id, written, withdrawn = import_one(
                cursor, product, args.eur_inr_rate,
                recipe_for(recipes, product))
            db.commit()
            print("  imported %-14s product_id=%s  %d %s row(s)%s"
                  % (product["source_ref"], product_id, written,
                     product["param_mode"].lower(),
                     ", %d withdrawn" % withdrawn if withdrawn else ""))
        except Exception as exc:                # noqa: BLE001 - reported, not hidden
            db.rollback()
            failures.append((product["source_ref"], exc))
            print("  FAILED   %-14s rolled back: %s"
                  % (product["source_ref"], exc))
    print("\n%d imported, %d failed. merchant_enabled stays 0 — release is a "
          "separate decision." % (len(products) - len(failures), len(failures)))
    if failures:
        sys.exit(1)


if __name__ == "__main__":
    main()
